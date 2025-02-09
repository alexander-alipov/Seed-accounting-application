import tkinter as tk
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Alignment
import os

path = os.getcwd()+'\\'

"ЗАПОЛНЕНИЕ ЯЧЕЕК EXCEL"
def Data():
    culture = ws['B'][indx].value
    variety = ws['C'][indx].value
    if ws['H'][indx].value == None:
        remain = int(ws["G"][indx].value)
        ws['H'][indx].value = given_seeds
        ws['I'][indx].value = remain - given_seeds
    else:
        remain = int(ws['G'][indx].value)
        ws['H'][indx].value = int(ws['H'][indx].value) + given_seeds
        ws['I'][indx].value = int(ws['I'][indx].value) - given_seeds
    ws['J'][indx].value = date
    ws['H'][indx].alignment = Alignment(horizontal='center')
    ws['I'][indx].alignment = Alignment(horizontal='center')
    ws['J'][indx].alignment = Alignment(horizontal='center')
    if int(remain) - given_seeds < 0:
        root2 = tk.Tk()
        root2.withdraw()
        if not messagebox.askokcancel("Отрицательное значение семян", \
                                     f"На складе осталось {remain} семян, а выдается {given_seeds}. Нажмите 'ок', чтобы внести изменения или выберите отмену"):
            root2.destroy()
            return None
        root2.destroy()
    return [culture, variety]

"ЗАПОЛНЕНИЕ СТРАНИЦЫ ЛОГИРОВАНИЕ ЗАПИСЕЙ"
def logging_data(number, given_seeds, who_gave, receiver, date, aim):
    ws2 = wb["Логирование записей"]
    ws2.append([number, given_seeds, who_gave, receiver, date, aim])
    ind = ws2.max_row
    ws2['A'][ind-1].alignment = Alignment(horizontal='center')
    ws2['B'][ind-1].alignment = Alignment(horizontal='center')
    ws2['C'][ind-1].alignment = Alignment(horizontal='center')
    ws2['D'][ind-1].alignment = Alignment(horizontal='center')
    ws2['E'][ind-1].alignment = Alignment(horizontal='center')
    ws2['F'][ind-1].alignment = Alignment(horizontal='center')

"ОТПРАВКА ДАННЫХ НА ДИСК В БИТРИКС"
def send_to_disk():
    global bitrix
    bitrix = True
    wb.save(path + 'Список гибридов с описанием.xlsx')
    del wb['Описание гибридов'], wb['Болезни растений'], wb['Пересчеты значений'], wb['Report']
    wb.save(path + 'реестр семян.xlsx')
    os.popen(path+"bitrix_upload.exe")
    Exit()
    
"ОБРАБОТЧИК КНОПКИ ДОБАВИТЬ"    
def getInput():
    global params
    number_val.get()
    params = (number.get(), given_seeds.get(), who_gave.get(), receiver.get(), date.get(), combo_aim.get())
    root.destroy()

"ЗАКРЫТИЕ ДИАЛОГОВОГО ОКНА"
def Exit():
    global flag
    flag = True
    root.destroy()

"ОБРАБОТКА КОПКИ НОМЕРА ОБРАЗЦА"
def processing_number(number):
    result = processing_number_val(number)
    if not result:
        root2 = tk.Tk()
        root2.withdraw()
        messagebox.showerror("Ошибка", 'Номер образца не существует или неверный формат номера образца')
        root2.destroy()
    else:
        return (result[0], result[1])

"ОБРАБОТКА НОМЕРА number_val = tk.StringVar()"
def processing_number_val(number):
    def check(number2):
        if number2 in tp:
            return (number2, tp.index(number2))
        else:
            return False
    if '№' in number:
        return check(number)
    else:
        return check('№' + number)

def check_number(*args):
    try:
        num, indx = processing_number_val(number_val.get())
        result_number.set(f'{num} {ws["B"][indx].value} {ws["C"][indx].value}')
    except:
        try:
            result_number.set(f'{num} отсутствует')
        except:
            pass
def check_remain(*args):
    try:
        num, indx = processing_number_val(number_val.get())
        current_remain = int(ws["G"][indx].value) if ws["H"][indx].value == None else ws["I"][indx].value
        result_remain.set(f'{num} всего осталось {current_remain} семян')
    except:
        result_remain.set(f'{num} - данные по семенам отсутствуют')

wb = openpyxl.load_workbook(path + 'Список гибридов с описанием.xlsx')
ws = wb["Учет семян"]
tp = tuple(ws.iter_cols(max_col=1, values_only=True))[0]
d, bitrix = None, False

while True:
    flag = False
    root = tk.Tk()
    root.title("Окно внесения данных")
    root.geometry('650x550')
    Font_ = ("Times New Roman", 17)

    number_val = tk.StringVar()
    result_number = tk.StringVar()
    remain_val = tk.StringVar()
    result_remain = tk.StringVar()

    tk.Label(root, textvariable = result_number).grid(row = 0, column = 2, sticky = 'W')
    tk.Label(root, text = "    Номер образца №: ", font=Font_).grid(row = 0, column = 0, sticky = 'W')
    tk.Label(root, textvariable = result_remain).grid(row=1, column=2, sticky='W')
    tk.Label(root, text = "    Выдано семян:",font=Font_).grid(row = 1, column = 0, sticky = 'W')
    tk.Label(root, text = "    Выдал(а):",font=Font_).grid(row = 2, column = 0, sticky = 'W')
    tk.Label(root, text = "    Получатель:",font=Font_).grid(row = 3, column = 0, sticky = 'W')
    tk.Label(root, text = "    Дата выдачи:",font=Font_).grid(row = 4, column = 0, sticky = 'W')
    tk.Label(root, text = "    Цель использования:",font=Font_).grid(row = 5, column = 0, sticky = 'W')

    number = tk.Entry(root, textvariable=number_val)
    number.grid(row = 0, column = 1, pady=0)
    given_seeds = tk.Entry(root, textvariable=remain_val)
    given_seeds.grid(row = 1, column = 1, pady=0)
    who_gave = tk.Entry(root)
    who_gave.grid(row = 2, column = 1,  pady=0)
    receiver = tk.Entry(root)
    receiver.grid(row = 3, column = 1, pady=0)
    date = tk.Entry(root)
    date.grid(row = 4, column = 1, pady=0)

    number_val.trace_add("write", check_number)
    remain_val.trace_add("write", check_remain)

    val = ('xxx', 'yyy', 'zzz', 'www')
    combo_aim = Combobox(root, values = val, width = 24)
    combo_aim.grid(row = 5, column = 1)

    tk.Button(root, text = "Выход", command = Exit, \
              font=Font_).grid(column=0, row = 6, pady=10, padx=20, sticky = 'W')
    tk.Button(root, text = "Добавить", command = getInput, \
              font=Font_).grid(column=1, row = 6, pady=10, padx=0, sticky = 'W')
    tk.Button(root, text = "Отправить в Битрикс", command = send_to_disk, \
              font=Font_).grid(column=0, row = 7, pady=0, padx=0, sticky = 'W')

    root.protocol("WM_DELETE_WINDOW", Exit)
    root.mainloop()
    
    if flag:
        break
    number, indx = processing_number(params[0])
    given_seeds = int(params[1])
    who_gave = params[2]
    receiver = params[3]
    date = params[4]
    aim = params[5]

    try:
        d = Data()
        if d != None:
            logging_data(number, given_seeds, who_gave, receiver, date, aim)
            root2 = tk.Tk()
            root2.withdraw()
            messagebox.showinfo('Данные успешно внесены', f'Образец {number} {d[0]} {d[1]} добавлен')
            root2.destroy()
    except PermissionError:
        root2 = tk.Tk()
        root2.withdraw()
        messagebox.showerror('Отказано в доступе', 'Закройте файл "Реестр семян" и повторите попытку')
        root2.destroy()
        break
    except TypeError:
        root2 = tk.Tk()
        root2.withdraw()
        messagebox.showerror('Ошибка внесения данных',f'В таблице не хватает данных. Проверьте ячейки образца {number}')
        root2.destroy()
        break

if d != None and bitrix == False:
    wb.save(path + 'Список гибридов с описанием.xlsx')





